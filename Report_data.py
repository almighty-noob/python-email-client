import openpyxl
Excel_file = openpyxl.load_workbook("D:\\javatpoint\\Class 12 NCERT\\Daily and Monthly Report.xlsx")
Current_Sheet = Excel_file.active

Date_and_time = Current_Sheet.cell(row = Current_Sheet.max_row, column = 1)
Topic = Current_Sheet.cell(row = Current_Sheet.max_row, column = 2)
Word_count = Current_Sheet.cell(row = Current_Sheet.max_row, column = 3)
Date = Date_and_time.value.strftime("%d-%m-%Y")
Final_Word_count = eval(Word_count.value[1:])
Final_Topic_name = Topic.value

# print(Date)